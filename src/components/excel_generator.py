"""
Generador de archivos Excel - PEF ToolBoard v2.0
Rellena la plantilla PEF_TOOLBOARD_v20.xlsx con los datos de session_state.
Las fórmulas del Excel quedan intactas y recalculan automáticamente al abrirlo.
"""

import io
import warnings
from pathlib import Path

from openpyxl import load_workbook

TEMPLATE_PATH = Path(__file__).parent.parent.parent / "templates" / "PEF_TOOLBOARD_v20.xlsx"

CAPEX_ROWS = {
    "investigacion": 26,
    "patentes": 27,
    "aplicaciones": 28,
    "otros_intangibles": 29,
    "terrenos": 31,
    "instalaciones": 32,
    "maquinaria": 33,
    "equipos": 34,
    "mobiliario": 35,
    "vehiculos": 36,
    "otros_materiales": 37,
    "fianzas": 40,
}

OPEX_ROWS = {
    "alquileres": 77,
    "suministros": 78,
    "rentings": 79,
    "reparaciones": 80,
    "servicios_prof": 81,
    "transportes": 82,
    "bancarios_seguros": 83,
    "marketing": 84,
    "tributos": 85,
}

NOMINAS_ROWS = {
    (1, "socios"): 90,
    (1, "perfil_a"): 91,
    (1, "perfil_b"): 92,
    (1, "perfil_c"): 93,
    (1, "perfil_d"): 94,
    (2, "socios"): 95,
    (2, "perfil_a"): 96,
    (2, "perfil_b"): 97,
    (2, "perfil_c"): 98,
    (2, "perfil_d"): 99,
    (3, "socios"): 100,
    (3, "perfil_a"): 101,
    (3, "perfil_b"): 102,
    (3, "perfil_c"): 103,
    (3, "perfil_d"): 104,
}


def read_template(excel_bytes: bytes) -> dict:
    """Lee un Excel PEF_TOOLBOARD generado y extrae los datos para session_state.

    Inverso exacto de fill_template: lee las celdas INPUT y reconstruye la
    estructura de session_state. Los valores porcentuales (SOM, CV, interés,
    incrementos) se multiplican ×100 al leer (estaban almacenados como decimal).

    Args:
        excel_bytes: Contenido del archivo Excel (bytes)

    Returns:
        dict: Claves y valores listos para asignar a st.session_state
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)

    ws_idea = wb["IDEA"]
    ws_hip = wb["HIPOTESIS "]
    ws_an = wb["ANALISIS "]

    def v(cell, default=0):
        """Valor numérico de celda, con default para None."""
        val = cell.value
        return val if val is not None else default

    def vs(cell):
        """Valor de celda como string."""
        val = cell.value
        return str(val) if val is not None else ""

    result = {}

    # --- IDEA ---
    result["proyecto"] = {
        "nombre": vs(ws_idea["C11"]),
        "equipo": vs(ws_idea["C12"]),
        "fecha_inicio": vs(ws_idea["C13"]),
    }

    # --- Fiscalidad ---
    result["fiscalidad"] = {
        "is_rate":           v(ws_hip["C9"], 0.25) * 100,
        "iva_compras":       v(ws_hip["C10"], 0.21) * 100,
        "iva_ventas":        v(ws_hip["C11"], 0.21) * 100,
        "iva_inversiones":   v(ws_hip["C12"], 0.21) * 100,
        "ss_autonomos_rate": v(ws_hip["C13"], 0.15) * 100,
        "ss_tope_autonomos": v(ws_hip["E13"], 56640.0),
        "ss_empresa_rate":   v(ws_hip["C14"], 0.33) * 100,
        "ss_trabajador_rate":v(ws_hip["D14"], 0.065) * 100,
        "ss_tope_general":   v(ws_hip["E14"], 56640.0),
        "irpf_bajo":         v(ws_hip["C15"], 0.0) * 100,
        "irpf_medio":        v(ws_hip["D15"], 0.20) * 100,
        "irpf_alto":         v(ws_hip["E15"], 0.40) * 100,
    }

    # --- CAPEX inicial (subvención no está en la plantilla; queda a 0) ---
    capex = {}
    for key, row in CAPEX_ROWS.items():
        capex[key] = {
            "importe": int(v(ws_hip[f"C{row}"])),
            "anos": int(v(ws_hip[f"F{row}"])),
            "subvencion": 0,
        }
    result["capex"] = capex

    # --- Proyectos de inversión ---
    result["proyectos_inversion"] = {
        "proyecto_inv_1": {
            "importe": int(v(ws_hip["C44"])),
            "anos": int(v(ws_hip["F44"])),
            "mes_adquisicion": int(v(ws_hip["G44"], 13)),
            "subvencion": int(v(ws_hip["C45"])),
            "observaciones": "",
        },
        "proyecto_inv_2": {
            "importe": int(v(ws_hip["C47"])),
            "anos": int(v(ws_hip["F47"])),
            "mes_adquisicion": int(v(ws_hip["G47"], 13)),
            "subvencion": int(v(ws_hip["C48"])),
            "observaciones": "",
        },
    }

    # --- Proyectos de trabajo activo propio ---
    result["proyectos_trabajo"] = {
        "proyecto_trab_1": {
            "importe": int(v(ws_hip["C52"])),
            "anos": int(v(ws_hip["F52"])),
            "mes_inicio": int(v(ws_hip["G52"], 1)),
            "mes_fin": int(v(ws_hip["H52"], 1)),
            "subvencion": int(v(ws_hip["C53"])),
            "observaciones": "",
        },
        "proyecto_trab_2": {
            "importe": int(v(ws_hip["C55"])),
            "anos": int(v(ws_hip["F55"])),
            "mes_inicio": int(v(ws_hip["G55"], 1)),
            "mes_fin": int(v(ws_hip["H55"], 1)),
            "subvencion": int(v(ws_hip["C56"])),
            "observaciones": "",
        },
    }

    # --- Financiación ---
    result["financiacion"] = {
        "capital_inicial": {
            "importe": int(v(ws_hip["D63"])),
            "acciones": int(v(ws_hip["F63"])),
        },
        "ampliacion": {
            "mes": int(v(ws_an["K8"], 21)),
            "importe": int(v(ws_an["L8"])),
            "valoracion_premoney": int(v(ws_an["M8"])),
        },
        "prestamo1": {
            "mes_inicio": int(v(ws_hip["C66"], 1)),
            "importe": int(v(ws_hip["D66"])),
            "meses_carencia": int(v(ws_hip["E66"])),
            "meses_amortizacion": int(v(ws_hip["F66"], 60)),
            "interes": float(v(ws_hip["H66"])) * 100,  # decimal → %
        },
        "prestamo2": {
            "mes_inicio": int(v(ws_hip["C67"], 1)),
            "importe": int(v(ws_hip["D67"])),
            "meses_carencia": int(v(ws_hip["E67"])),
            "meses_amortizacion": int(v(ws_hip["F67"], 60)),
            "interes": float(v(ws_hip["H67"])) * 100,
        },
        "poliza_interes": float(v(ws_hip["D70"])) * 100,
    }

    # --- OPEX gastos fijos ---
    # E=año2 (siempre valor), H=año3, K=año4, N=año5 (pueden ser fórmulas
    # si el usuario no las editó en Excel). Si son fórmulas (str), se hereda
    # el valor del año anterior para mantener coherencia en session_state.
    gastos_fijos = {}
    for key, row in OPEX_ROWS.items():
        def _read_inc(col, fallback):
            raw = ws_hip[f"{col}{row}"].value
            if isinstance(raw, (int, float)):
                return float(raw) * 100
            return fallback  # celda aún es fórmula → heredar año anterior

        inc2 = float(v(ws_hip[f"E{row}"])) * 100
        inc3 = _read_inc("H", inc2)
        inc4 = _read_inc("K", inc3)
        inc5 = _read_inc("N", inc4)
        gastos_fijos[key] = {
            "ano1": float(v(ws_hip[f"C{row}"])),
            "incrementos": [inc2, inc3, inc4, inc5],
        }
    result["opex"] = {"gastos_fijos": gastos_fijos, "empleados": []}

    # --- Nóminas (empleados_data) ---
    etapa_socios_rows = {1: 90, 2: 95, 3: 100}
    emp_data = {}
    for etapa in [1, 2, 3]:
        socios_row = etapa_socios_rows[etapa]
        default_inc = 0.02 if etapa < 3 else 0.03
        inc_raw = v(ws_hip[f"N{socios_row}"], default_inc)
        emp_data[etapa] = {
            "incremento_salario": float(inc_raw) * 100,
            "perfiles": {},
        }
    for (etapa, key), row in NOMINAS_ROWS.items():
        emp_data[etapa]["perfiles"][key] = {
            "num": int(v(ws_hip[f"D{row}"])),
            "alta": int(v(ws_hip[f"E{row}"], 1)),
            "baja": int(v(ws_hip[f"F{row}"], 60)),
            "salario": v(ws_hip[f"G{row}"]),
        }
    result["empleados_data"] = emp_data

    # --- Ventas / Ingresos ---
    tipo_cfg = [
        ("tipo_a", 110, 111, 122, 128, 129, 130),
        ("tipo_b", 113, 114, 123, 132, 133, 134),
        ("tipo_c", 116, 117, 124, 136, 137, 138),
    ]
    cols_som = ["C", "D", "E", "F", "G"]
    incremento_precios = [v(ws_hip[f"{c}125"]) * 100 for c in ["D", "E", "F", "G"]]
    nombres = {"tipo_a": "Producto A", "tipo_b": "Producto B", "tipo_c": "Producto C"}

    ingresos = {}
    for tipo, row_sam, row_som, row_precio, row_cv_prod, row_cv_acq, row_comis in tipo_cfg:
        som = [float(v(ws_hip[f"{col}{row_som}"])) * 100 for col in cols_som]
        ingresos[tipo] = {
            "nombre": nombres[tipo],
            "sam": int(v(ws_hip[f"C{row_sam}"])),
            "som": som,
            "precio": float(v(ws_hip[f"C{row_precio}"])),
            "incremento": incremento_precios,
            "cv_produccion": float(v(ws_hip[f"C{row_cv_prod}"])) * 100,
            "cv_adquisicion": float(v(ws_hip[f"C{row_cv_acq}"])) * 100,
            "comisiones": float(v(ws_hip[f"C{row_comis}"])) * 100,
        }
    result["ingresos"] = ingresos

    return result


def fill_template(ss) -> bytes:
    """Carga la plantilla PEF_TOOLBOARD y escribe los datos de session_state.

    Args:
        ss: st.session_state (o cualquier dict-like con los datos del proyecto)

    Returns:
        bytes: Contenido del archivo Excel generado
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(TEMPLATE_PATH)

    ws_idea = wb["IDEA"]
    ws_hip = wb["HIPOTESIS "]   # trailing space
    ws_an = wb["ANALISIS "]    # trailing space

    _fill_idea(ws_idea, ss)
    _fill_fiscal(ws_hip, ss)
    _fill_capex(ws_hip, ss)
    _fill_financiacion(ws_hip, ws_an, ss)
    _fill_opex(ws_hip, ss)
    _fill_nominas(ws_hip, ss)
    _fill_ventas(ws_hip, ss)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Helpers de relleno por sección
# ---------------------------------------------------------------------------

def _fill_idea(ws, ss):
    """Rellena la hoja IDEA con los datos del proyecto."""
    proyecto = ss.get("proyecto", {})
    ws["C11"] = proyecto.get("nombre", "")
    ws["C12"] = proyecto.get("equipo", "")
    ws["C13"] = proyecto.get("fecha_inicio", "")


def _fill_fiscal(ws, ss):
    """Escribe los parámetros fiscales en HIPOTESIS desde session_state."""
    f = ss.get("fiscalidad", {})
    ws["C9"]  = f.get("is_rate", 25.0) / 100
    ws["C10"] = f.get("iva_compras", 21.0) / 100
    ws["C11"] = f.get("iva_ventas", 21.0) / 100
    ws["C12"] = f.get("iva_inversiones", 21.0) / 100
    ws["C13"] = f.get("ss_autonomos_rate", 15.0) / 100
    ws["D13"] = 0.0  # SS autónomos trabajador (siempre 0 en régimen especial)
    ws["E13"] = f.get("ss_tope_autonomos", 56640.0)
    ws["C14"] = f.get("ss_empresa_rate", 33.0) / 100
    ws["D14"] = f.get("ss_trabajador_rate", 6.5) / 100
    ws["E14"] = f.get("ss_tope_general", 56640.0)
    ws["C15"] = f.get("irpf_bajo", 0.0) / 100
    ws["D15"] = f.get("irpf_medio", 20.0) / 100
    ws["E15"] = f.get("irpf_alto", 40.0) / 100


def _fill_capex(ws, ss):
    """Rellena CAPEX inicial, proyectos de inversión y proyectos de trabajo."""
    capex = ss.get("capex", {})

    # CAPEX inicial (12 categorías, mes adquisición siempre = 1)
    for key, row in CAPEX_ROWS.items():
        data = capex.get(key, {})
        ws[f"C{row}"] = data.get("importe", 0)
        ws[f"F{row}"] = data.get("anos", 0)
        ws[f"G{row}"] = 1

    # Proyectos de inversión en años posteriores
    proyectos_inv = ss.get("proyectos_inversion", {})

    p1 = proyectos_inv.get("proyecto_inv_1", {})
    ws["C44"] = p1.get("importe", 0)
    ws["F44"] = p1.get("anos", 0)
    ws["G44"] = p1.get("mes_adquisicion", 13)
    ws["C45"] = p1.get("subvencion", 0)
    ws["G45"] = p1.get("mes_adquisicion", 13)  # inicio_subv_inv1

    # Si proyecto_inv_1 no tiene subvención propia, usar el slot para subvenciones de CAPEX inicial
    total_capex_subvencion = sum(data.get("subvencion", 0) for data in capex.values())
    if total_capex_subvencion > 0 and p1.get("importe", 0) == 0 and p1.get("subvencion", 0) == 0:
        ws["C45"] = total_capex_subvencion
        ws["G45"] = 1  # CAPEX inicial: subvención recibida en mes 1

    p2 = proyectos_inv.get("proyecto_inv_2", {})
    ws["C47"] = p2.get("importe", 0)
    ws["F47"] = p2.get("anos", 0)
    ws["G47"] = p2.get("mes_adquisicion", 13)
    ws["C48"] = p2.get("subvencion", 0)
    ws["G48"] = p2.get("mes_adquisicion", 13)  # inicio_subv_inv2

    # Proyectos de trabajo para activo propio
    proyectos_trab = ss.get("proyectos_trabajo", {})

    t1 = proyectos_trab.get("proyecto_trab_1", {})
    ws["C52"] = t1.get("importe", 0)
    ws["F52"] = t1.get("anos", 0)
    ws["G52"] = t1.get("mes_inicio", 1)
    ws["H52"] = t1.get("mes_fin", 1)
    ws["C53"] = t1.get("subvencion", 0)

    t2 = proyectos_trab.get("proyecto_trab_2", {})
    ws["C55"] = t2.get("importe", 0)
    ws["F55"] = t2.get("anos", 0)
    ws["G55"] = t2.get("mes_inicio", 1)
    ws["H55"] = t2.get("mes_fin", 1)
    ws["C56"] = t2.get("subvencion", 0)


def _fill_financiacion(ws_hip, ws_an, ss):
    """Rellena financiación: capital, préstamos, póliza y ampliación."""
    fin = ss.get("financiacion", {})

    # Capital inicial (mes siempre 1)
    cap = fin.get("capital_inicial", {})
    ws_hip["C63"] = 1
    ws_hip["D63"] = cap.get("importe", 0)
    ws_hip["F63"] = cap.get("acciones", 0)

    # Préstamo 1
    p1 = fin.get("prestamo1", {})
    ws_hip["C66"] = p1.get("mes_inicio", 1)
    ws_hip["D66"] = p1.get("importe", 0)
    ws_hip["E66"] = p1.get("meses_carencia", 0)
    ws_hip["F66"] = p1.get("meses_amortizacion", 60)
    ws_hip["H66"] = p1.get("interes", 0) / 100  # almacenado como % → decimal

    # Préstamo 2
    p2 = fin.get("prestamo2", {})
    ws_hip["C67"] = p2.get("mes_inicio", 1)
    ws_hip["D67"] = p2.get("importe", 0)
    ws_hip["E67"] = p2.get("meses_carencia", 0)
    ws_hip["F67"] = p2.get("meses_amortizacion", 60)
    ws_hip["H67"] = p2.get("interes", 0) / 100

    # Póliza de crédito (solo interés; el importe es ilimitado en la plantilla)
    ws_hip["D70"] = fin.get("poliza_interes", 0) / 100

    # Ampliación de capital (hoja ANALISIS)
    amp = fin.get("ampliacion", {})
    ws_an["K8"] = amp.get("mes", 0)
    ws_an["L8"] = amp.get("importe", 0)
    ws_an["M8"] = amp.get("valoracion_premoney", 0)


def _fill_opex(ws, ss):
    """Rellena gastos fijos OPEX: importe año 1 e incrementos años 2-5.

    Escribe las 4 celdas de incremento (E, H, K, N) explícitamente para
    preservar valores independientes por año, sobreescribiendo las fórmulas
    de la plantilla (=E, =H, =K).
    """
    opex = ss.get("opex", {})
    gastos_fijos = opex.get("gastos_fijos", {})

    # E=año2, H=año3, K=año4, N=año5
    inc_cols = ["E", "H", "K", "N"]

    for key, row in OPEX_ROWS.items():
        data = gastos_fijos.get(key, {})
        ws[f"C{row}"] = data.get("ano1", 0)
        incrementos = data.get("incrementos", [0.0, 0.0, 0.0, 0.0])
        for i, col in enumerate(inc_cols):
            val = incrementos[i] if i < len(incrementos) else 0.0
            ws[f"{col}{row}"] = val / 100  # % → decimal


def _fill_nominas(ws, ss):
    """Rellena la tabla de nóminas (3 etapas × 5 perfiles)."""
    emp_data = ss.get("empleados_data", {})
    etapa_socios_rows = {1: 90, 2: 95, 3: 100}

    for (etapa, key), row in NOMINAS_ROWS.items():
        etapa_data = emp_data.get(etapa, {})
        perfiles = etapa_data.get("perfiles", {})
        perfil = perfiles.get(key, {})
        ws[f"D{row}"] = perfil.get("num", 0)
        ws[f"E{row}"] = perfil.get("alta", 1)
        ws[f"F{row}"] = perfil.get("baja", 60)
        ws[f"G{row}"] = perfil.get("salario", 0)

    for etapa, socios_row in etapa_socios_rows.items():
        etapa_data = emp_data.get(etapa, {})
        inc_pct = etapa_data.get("incremento_salario", 2.0 if etapa < 3 else 3.0)
        ws[f"N{socios_row}"] = inc_pct / 100


def _fill_ventas(ws, ss):
    """Rellena SAM, SOM, precios, incrementos y costes variables."""
    ingresos = ss.get("ingresos", {})

    # (clave_ss, row_sam, row_som, row_precio, row_cv_prod, row_cv_acq, row_comis)
    tipo_cfg = [
        ("tipo_a", 110, 111, 122, 128, 129, 130),
        ("tipo_b", 113, 114, 123, 132, 133, 134),
        ("tipo_c", 116, 117, 124, 136, 137, 138),
    ]
    cols_som = ["C", "D", "E", "F", "G"]  # 5 años

    for tipo, row_sam, row_som, row_precio, row_cv_prod, row_cv_acq, row_comis in tipo_cfg:
        data = ingresos.get(tipo, {})

        # SAM
        ws[f"C{row_sam}"] = data.get("sam", 0)

        # SOM % (almacenado como %, se escribe como decimal)
        som = data.get("som", [0, 0, 0, 0, 0])
        for i, col in enumerate(cols_som):
            val = som[i] if i < len(som) else 0
            ws[f"{col}{row_som}"] = val / 100

        # Precio año 1
        ws[f"C{row_precio}"] = data.get("precio", 0)

        # Costes variables (almacenados como %, se escriben como decimal)
        ws[f"C{row_cv_prod}"] = data.get("cv_produccion", 0) / 100
        ws[f"C{row_cv_acq}"] = data.get("cv_adquisicion", 0) / 100
        ws[f"C{row_comis}"] = data.get("comisiones", 0) / 100

    # Incremento de precios años 2-5 (D125:G125), shared entre tipos
    # Se usa tipo_a como referencia
    incremento = ingresos.get("tipo_a", {}).get("incremento", [0, 0, 0, 0])
    for i, col in enumerate(["D", "E", "F", "G"]):
        val = incremento[i] if i < len(incremento) else 0
        ws[f"{col}125"] = val / 100
